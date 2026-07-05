from __future__ import annotations

import csv
import io
import json
from math import pi
from collections import OrderedDict
from datetime import datetime
from typing import Iterable
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func
from sqlalchemy.orm import Session

from .config import get_settings
from .models import AppSetting, Asset, Finding, ScanJob

SEVERITY_ORDER = ["Critical", "High", "Medium", "Low", "Informational"]
STATUS_ORDER = ["Open", "Assigned", "In Progress", "Remediated", "Retest", "Closed", "False Positive", "Accepted Risk"]
SCANNER_ORDER = ["trivy", "syft", "grype", "zap", "openvas"]


def _setting_value(db: Session, key: str) -> str | None:
    return db.query(AppSetting.value).filter(AppSetting.key == key).scalar()


def _normalize_gate_decision(value: str | None, fallback: str) -> str:
    normalized = (value or "").strip().lower()
    if normalized in {"allowed", "need_approval", "blocked", "pending"}:
        return normalized
    return fallback


def _parse_setting_bool(value: str | None, fallback: bool) -> bool:
    if value is None:
        return fallback
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _parse_setting_int(value: str | None, fallback: int) -> int:
    try:
        return max(0, int((value or "").strip()))
    except (TypeError, ValueError):
        return fallback


def _gate_runtime_config(db: Session) -> dict[str, object]:
    settings = get_settings()
    return {
        "gate_block_on_critical": _parse_setting_bool(
            _setting_value(db, "gate_block_on_critical"),
            settings.gate_block_on_critical,
        ),
        "gate_high_threshold": _parse_setting_int(
            _setting_value(db, "gate_high_threshold"),
            settings.gate_high_threshold,
        ),
        "gate_high_decision": _normalize_gate_decision(
            _setting_value(db, "gate_high_decision"),
            _normalize_gate_decision(settings.gate_high_decision, "need_approval"),
        ),
        "gate_medium_threshold": _parse_setting_int(
            _setting_value(db, "gate_medium_threshold"),
            settings.gate_medium_threshold,
        ),
        "gate_medium_decision": _normalize_gate_decision(
            _setting_value(db, "gate_medium_decision"),
            _normalize_gate_decision(settings.gate_medium_decision, "allowed"),
        ),
        "gate_low_threshold": _parse_setting_int(
            _setting_value(db, "gate_low_threshold"),
            settings.gate_low_threshold,
        ),
        "gate_low_decision": _normalize_gate_decision(
            _setting_value(db, "gate_low_decision"),
            _normalize_gate_decision(settings.gate_low_decision, "allowed"),
        ),
    }


def _release_metadata(scan: ScanJob) -> dict[str, object]:
    if not scan.release or not scan.release.metadata_json:
        return {}
    try:
        parsed = json.loads(scan.release.metadata_json)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        return {}
    return {}


def _severity_summary_for_scan(db: Session, scan_id: int) -> dict[str, int]:
    rows = (
        db.query(Finding.severity_normalized, func.count(Finding.id))
        .filter(Finding.scan_job_id == scan_id)
        .group_by(Finding.severity_normalized)
        .all()
    )
    summary = {"critical": 0, "high": 0, "medium": 0, "low": 0, "informational": 0, "total": 0}
    for severity, count in rows:
        normalized = (severity or "").strip().lower()
        if normalized == "critical":
            summary["critical"] += int(count)
        elif normalized == "high":
            summary["high"] += int(count)
        elif normalized == "medium":
            summary["medium"] += int(count)
        elif normalized == "low":
            summary["low"] += int(count)
        else:
            summary["informational"] += int(count)
        summary["total"] += int(count)
    return summary


def _gate_decision_for_scan(scan: ScanJob, db: Session) -> str:
    gate_config = _gate_runtime_config(db)
    severity_summary = _severity_summary_for_scan(db, scan.id)
    if scan.status in {"queued", "running"}:
        return "pending"
    if scan.status != "completed":
        return "blocked"
    release_metadata = _release_metadata(scan)
    override_decision = _normalize_gate_decision(str(release_metadata.get("gate_override_decision") or ""), "")
    if override_decision:
        return override_decision
    if gate_config["gate_block_on_critical"] and severity_summary.get("critical", 0) > 0:
        return "blocked"
    if gate_config["gate_high_threshold"] > 0 and severity_summary.get("high", 0) >= gate_config["gate_high_threshold"]:
        return _normalize_gate_decision(str(gate_config["gate_high_decision"]), "need_approval")
    if gate_config["gate_medium_threshold"] > 0 and severity_summary.get("medium", 0) >= gate_config["gate_medium_threshold"]:
        return _normalize_gate_decision(str(gate_config["gate_medium_decision"]), "allowed")
    if gate_config["gate_low_threshold"] > 0 and severity_summary.get("low", 0) >= gate_config["gate_low_threshold"]:
        return _normalize_gate_decision(str(gate_config["gate_low_decision"]), "allowed")
    return "allowed"


def _ordered_counts(rows: Iterable[tuple[str | None, int]], order: list[str]) -> OrderedDict[str, int]:
    data = {k or "Unknown": int(v) for k, v in rows}
    result: OrderedDict[str, int] = OrderedDict()
    for key in order:
        result[key] = data.pop(key, 0)
    for key in sorted(data):
        result[key] = data[key]
    return result


def active_findings_query(db: Session):
    return (
        db.query(Finding)
        .join(Asset, Finding.asset_id == Asset.id)
        .join(ScanJob, Finding.scan_job_id == ScanJob.id)
        .filter(Asset.is_active == True, ScanJob.is_visible == True)  # noqa: E712
    )


def get_summary(db: Session) -> dict:
    active_findings = active_findings_query(db)
    severity = _ordered_counts(
        active_findings.with_entities(Finding.severity_normalized, func.count(Finding.id)).group_by(Finding.severity_normalized).all(),
        SEVERITY_ORDER,
    )
    status = _ordered_counts(
        active_findings.with_entities(Finding.status, func.count(Finding.id)).group_by(Finding.status).all(),
        STATUS_ORDER,
    )
    scanner = _ordered_counts(
        active_findings.with_entities(Finding.scanner, func.count(Finding.id)).group_by(Finding.scanner).all(),
        SCANNER_ORDER,
    )
    total_findings = sum(severity.values())
    open_findings = sum(v for k, v in status.items() if k not in {"Closed", "False Positive", "Accepted Risk"})
    critical_high = severity.get("Critical", 0) + severity.get("High", 0)
    completed_scans = db.query(ScanJob).filter(ScanJob.is_visible == True, ScanJob.status == "completed").count()  # noqa: E712
    failed_scans = db.query(ScanJob).filter(ScanJob.is_visible == True, ScanJob.status == "failed").count()  # noqa: E712

    top_assets = (
        db.query(Asset.id, Asset.name, func.count(Finding.id).label("total"), func.max(Finding.risk_score).label("max_risk"))
        .join(Finding, Finding.asset_id == Asset.id)
        .join(ScanJob, Finding.scan_job_id == ScanJob.id)
        .filter(Asset.is_active == True, ScanJob.is_visible == True)  # noqa: E712
        .group_by(Asset.id, Asset.name)
        .order_by(func.max(Finding.risk_score).desc(), func.count(Finding.id).desc())
        .limit(10)
        .all()
    )
    recent_scans = db.query(ScanJob).filter(ScanJob.is_visible == True).order_by(ScanJob.id.desc()).limit(20).all()  # noqa: E712
    latest_findings = active_findings.order_by(Finding.risk_score.desc(), Finding.id.desc()).limit(20).all()

    top_asset_rows = []
    for asset_id, name, total, max_risk in top_assets:
        latest_scan = (
            db.query(ScanJob)
            .filter(ScanJob.asset_id == asset_id, ScanJob.is_visible == True)  # noqa: E712
            .order_by(ScanJob.id.desc())
            .first()
        )
        gate_decision = _gate_decision_for_scan(latest_scan, db) if latest_scan else "allowed"
        top_asset_rows.append({
            "name": name,
            "total": total,
            "max_risk": max_risk,
            "gate_decision": gate_decision,
        })

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "assets": db.query(Asset).filter(Asset.is_active == True).count(),  # noqa: E712
        "scans": db.query(ScanJob).filter(ScanJob.is_visible == True).count(),  # noqa: E712
        "completed_scans": completed_scans,
        "failed_scans": failed_scans,
        "findings": total_findings,
        "open_findings": open_findings,
        "critical_high": critical_high,
        "severity": severity,
        "status": status,
        "scanner": scanner,
        "top_assets": top_asset_rows,
        "recent_scans": recent_scans,
        "latest_findings": latest_findings,
    }


def severity_pie_style(severity: dict[str, int]) -> str:
    return count_pie_style(
        severity,
        SEVERITY_ORDER,
        {
            "Critical": "#ef4444",
            "High": "#f97316",
            "Medium": "#facc15",
            "Low": "#3b82f6",
            "Informational": "#94a3b8",
        },
    )


def count_pie_segments(
    counts: dict[str, int],
    order: list[str],
    colors: dict[str, str],
    href_builder,
) -> list[dict[str, str | int | float]]:
    total = sum(counts.values())
    radius = 42.0
    circumference = 2 * pi * radius
    if total <= 0:
        return []

    segments: list[dict[str, str | int | float]] = []
    drawn = 0.0
    for key in order:
        value = counts.get(key, 0)
        if value <= 0:
            continue
        portion = value / total
        dash_length = portion * circumference
        segments.append({
            "label": key,
            "value": value,
            "color": colors.get(key, "#64748b"),
            "dasharray": f"{dash_length:.3f} {circumference:.3f}",
            "dashoffset": f"{-drawn:.3f}",
            "href": href_builder(key),
        })
        drawn += dash_length
    return segments


def severity_pie_segments(severity: dict[str, int]) -> list[dict[str, str | int | float]]:
    colors = {
        "Critical": "#ef4444",
        "High": "#f97316",
        "Medium": "#facc15",
        "Low": "#3b82f6",
        "Informational": "#94a3b8",
    }
    return count_pie_segments(
        severity,
        SEVERITY_ORDER,
        colors,
        lambda key: f"/findings?severity={quote_plus(key)}",
    )


def count_pie_style(counts: dict[str, int], order: list[str], colors: dict[str, str]) -> str:
    total = max(1, sum(counts.values()))
    cursor = 0.0
    stops: list[str] = []
    for key in order:
        value = counts.get(key, 0)
        if value <= 0:
            continue
        start = cursor
        cursor += (value / total) * 100
        color = colors.get(key, "#64748b")
        stops.append(f"{color} {start:.2f}% {cursor:.2f}%")
    if not stops:
        stops.append("#334155 0% 100%")
    return "background: conic-gradient(" + ", ".join(stops) + ");"


def findings_as_rows(findings: list[Finding]) -> list[dict[str, str | int | None]]:
    rows = []
    for f in findings:
        rows.append({
            "finding_id": f.id,
            "asset": f.asset.name if f.asset else "-",
            "environment": f.asset.environment if f.asset else "-",
            "scanner": f.scanner,
            "type": f.finding_type,
            "severity": f.severity_normalized,
            "risk_score": f.risk_score,
            "title": f.title,
            "cve": f.cve,
            "cwe": f.cwe,
            "package": f.package_name,
            "installed_version": f.installed_version,
            "fixed_version": f.fixed_version,
            "affected_component": f.affected_component,
            "status": f.status,
            "iris_alert_id": f.iris_alert_id,
            "recommendation": f.recommendation,
        })
    return rows


def export_findings_csv(findings: list[Finding]) -> str:
    rows = findings_as_rows(findings)
    output = io.StringIO()
    if not rows:
        output.write("finding_id,asset,severity,risk_score,title,status\n")
        return output.getvalue()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def export_findings_xlsx(findings: list[Finding]) -> bytes:
    rows = findings_as_rows(findings)
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    headers = list(rows[0].keys()) if rows else ["finding_id", "asset", "severity", "risk_score", "title", "status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 55)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def executive_markdown_report(summary: dict) -> str:
    sev = summary["severity"]
    status = summary["status"]
    lines = [
        "# SATRIA Vulnerability Summary Report",
        "",
        f"Generated at: {summary['generated_at']}",
        "",
        "## Executive Summary",
        "",
        f"- Assets monitored: {summary['assets']}",
        f"- Total scan jobs: {summary['scans']}",
        f"- Completed scans: {summary['completed_scans']}",
        f"- Failed scans: {summary['failed_scans']}",
        f"- Total findings: {summary['findings']}",
        f"- Open/actionable findings: {summary['open_findings']}",
        f"- Critical + High findings: {summary['critical_high']}",
        "",
        "## Severity Breakdown",
        "",
    ]
    for key, value in sev.items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Status Breakdown", ""])
    for key, value in status.items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Top Risk Assets", ""])
    for item in summary["top_assets"]:
        lines.append(
            f"- {item['name']}: {item['total']} finding(s), max risk {item['max_risk']}, gate {item['gate_decision']}"
        )
    lines.extend(["", "## Recommended Follow-up", "", "- Validate Critical/High findings and send confirmed items to DFIR-IRIS as alerts/cases.", "- Assign remediation owners and perform retest after fixes.", "- Keep raw scanner reports as evidence for audit and closure."])
    return "\n".join(lines) + "\n"
