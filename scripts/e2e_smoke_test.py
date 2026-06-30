"""End-to-end smoke test for SATRIA MVP.

Runs locally without Docker daemon, PostgreSQL, Redis, or real scanner binaries.
It verifies the full application path in demo mode:
asset intake -> scan orchestration -> scanner reports -> normalization -> dashboard
vulnerability summary with severity pie -> export reports -> DFIR-IRIS ticket stub
-> remediation status update -> audit/log/report existence.
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)

os.environ.setdefault("DATABASE_URL", "sqlite:////tmp/satria_e2e_smoke.db")
os.environ.setdefault("REPORT_DIR", "/tmp/satria_e2e_reports")
os.environ.setdefault("SATRIA_DEMO_MODE", "true")
os.environ.setdefault("CELERY_BROKER_URL", "memory://")
os.environ.setdefault("CELERY_RESULT_BACKEND", "cache+memory://")
os.environ.setdefault("SCAN_TARGET_ALLOWLIST", "localhost,127.0.0.1,registry.local,github.local,app-dev.local,10.100.")
os.environ["IRIS_URL"] = ""
os.environ["IRIS_API_KEY"] = ""

def sqlite_path_from_url(url: str) -> Path | None:
    if not url.startswith("sqlite:///"):
        return None
    raw = url.replace("sqlite:///", "", 1)
    if raw.startswith("/"):
        return Path(raw)
    return Path(raw)


DB_PATH = sqlite_path_from_url(os.environ["DATABASE_URL"])
REPORT_DIR = Path(os.environ["REPORT_DIR"])
if DB_PATH is not None:
    DB_PATH.unlink(missing_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)
for old_report in REPORT_DIR.glob("*.json"):
    old_report.unlink()

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

import app.main as main  # noqa: E402
from app.database import SessionLocal, init_db  # noqa: E402
from app.models import Asset, AuditLog, Finding, ScanJob, TicketCase, TicketEvidence, TicketTask  # noqa: E402
from app.tasks import run_scan_job  # noqa: E402


def run_sync_delay(scan_job_id: int):
    return run_scan_job(scan_job_id)


main.run_scan_job.delay = run_sync_delay
init_db()

checks: list[str] = []
with TestClient(main.app) as client:
    health = client.get("/health")
    assert health.status_code == 200, health.text
    checks.append("01 health endpoint OK")

    asset_payloads = [
        {
            "name": "SATRIA Demo Container Image",
            "asset_type": "container_image",
            "target": "registry.local/sakti-api:latest",
            "environment": "development",
            "criticality": "critical",
            "owner": "SITP",
            "technical_pic": "petugas-ki",
            "profile": "full_container",
            "expected_scanners": {"trivy", "syft", "grype"},
        },
        {
            "name": "SATRIA Demo Source Repository",
            "asset_type": "source_repository",
            "target": "https://github.local/sitp/sakti-api",
            "environment": "development",
            "criticality": "high",
            "owner": "SITP",
            "technical_pic": "devsecops",
            "profile": "repo_security",
            "expected_scanners": {"trivy", "syft"},
        },
        {
            "name": "SATRIA Demo Internal Server",
            "asset_type": "server",
            "target": "10.100.244.90",
            "environment": "staging",
            "criticality": "high",
            "owner": "SITP",
            "technical_pic": "sysadmin",
            "profile": "infra_va",
            "expected_scanners": {"openvas"},
        },
        {
            "name": "SATRIA Demo Web Application",
            "asset_type": "web_application",
            "target": "https://app-dev.local",
            "environment": "staging",
            "criticality": "medium",
            "owner": "SITP",
            "technical_pic": "pentester",
            "profile": "web_baseline",
            "expected_scanners": {"zap"},
        },
    ]

    created_asset_ids: list[int] = []
    for payload in asset_payloads:
        profile = payload.pop("profile")
        expected_scanners = payload.pop("expected_scanners")
        resp = client.post("/api/assets", json=payload)
        assert resp.status_code == 200, resp.text
        asset_id = resp.json()["id"]
        created_asset_ids.append(asset_id)
        scan_resp = client.post("/api/scans", json={"asset_id": asset_id, "profile": profile})
        assert scan_resp.status_code == 200, scan_resp.text
        scanners = set(scan_resp.json()["scanner"].split("+"))
        assert expected_scanners.issubset(scanners), (profile, scanners, expected_scanners)
    checks.append("02 asset intake and scan job creation OK")

    with SessionLocal() as db:
        assert db.query(Asset).count() == 4
        assert db.query(ScanJob).count() == 4
        assert db.query(ScanJob).filter(ScanJob.status == "completed").count() == 4
        scanners_found = {row[0] for row in db.query(Finding.scanner).distinct().all()}
        assert {"trivy", "syft", "grype", "openvas", "zap"}.issubset(scanners_found), scanners_found
        assert db.query(Finding).count() >= 10
        assert db.query(Finding).filter(Finding.severity_normalized == "Critical").count() >= 1
        assert db.query(Finding).filter(Finding.severity_normalized == "High").count() >= 1
    checks.append("03 scanner worker, raw reports, normalizer, and DB persistence OK")

    for path in ["/", "/assets", "/scan/new", "/scans", "/findings", "/tickets", "/vulnerability-summary"]:
        page = client.get(path)
        assert page.status_code == 200, f"{path}: {page.status_code}\n{page.text[:300]}"
    summary_page = client.get("/vulnerability-summary")
    html = summary_page.text
    for token in ["Ringkasan lanjutan untuk reporting dan bulk action", "Risk Composition", "Finding Status Pie", "Scanner Source Pie", "Recent Scans", "Export Excel", "Kirim Tiket Critical/High", "conic-gradient"]:
        assert token in html, f"missing dashboard token: {token}"
    checks.append("04 dashboard pages and vulnerability summary severity pie OK")

    summary_api = client.get("/api/summary")
    assert summary_api.status_code == 200, summary_api.text
    summary = summary_api.json()
    assert summary["assets"] == 4, summary
    assert summary["completed_scans"] == 4, summary
    assert summary["findings"] >= 10, summary
    assert summary["severity"]["Critical"] >= 1, summary
    checks.append("05 summary API OK")

    csv_resp = client.get("/reports/findings.csv")
    assert csv_resp.status_code == 200, csv_resp.text[:200]
    assert "finding_id,asset" in csv_resp.text
    assert "SATRIA Demo" in csv_resp.text

    xlsx_resp = client.get("/reports/findings.xlsx")
    assert xlsx_resp.status_code == 200
    xlsx_path = Path("/tmp/satria_e2e_findings.xlsx")
    xlsx_path.write_bytes(xlsx_resp.content)
    wb = load_workbook(xlsx_path)
    assert "Findings" in wb.sheetnames
    assert wb["Findings"].max_row >= 2

    md_resp = client.get("/reports/executive.md")
    assert md_resp.status_code == 200
    assert "SATRIA Vulnerability Summary Report" in md_resp.text
    assert "Critical + High findings" in md_resp.text
    checks.append("06 CSV, Excel, and executive Markdown reporting OK")

    with SessionLocal() as db:
        finding = db.query(Finding).filter(Finding.severity_normalized.in_(["Critical", "High"])).order_by(Finding.risk_score.desc()).first()
        assert finding is not None
        single_resp = client.post(f"/findings/{finding.id}/send-to-iris", follow_redirects=False)
        assert single_resp.status_code == 303, single_resp.text
        db.refresh(finding)
        ticket = db.query(TicketCase).filter(TicketCase.finding_id == finding.id).first()
        assert ticket is not None
        assert ticket.remote_case_id and ticket.remote_case_id.startswith("IRIS-STUB-CASE-"), ticket.remote_case_id
        assert db.query(TicketTask).filter(TicketTask.ticket_case_id == ticket.id).count() >= 1
        assert db.query(TicketEvidence).filter(TicketEvidence.ticket_case_id == ticket.id).count() >= 1

    bulk_resp = client.post("/ticketing/send-critical-high-to-iris", follow_redirects=False)
    assert bulk_resp.status_code == 303, bulk_resp.text
    with SessionLocal() as db:
        actionable_high = db.query(Finding).filter(Finding.severity_normalized.in_(["Critical", "High"])).count()
        ticketed = db.query(TicketCase).join(Finding, TicketCase.finding_id == Finding.id).filter(
            Finding.severity_normalized.in_(["Critical", "High"])
        ).count()
        assert ticketed == actionable_high, (ticketed, actionable_high)
    checks.append("07 DFIR-IRIS single and bulk ticketing workflow OK")

    with SessionLocal() as db:
        finding = db.query(Finding).order_by(Finding.risk_score.desc()).first()
        assert finding is not None
        status_resp = client.post(f"/findings/{finding.id}/status", data={"status": "Closed"}, follow_redirects=False)
        assert status_resp.status_code == 303, status_resp.text
        db.refresh(finding)
        assert finding.status == "Closed"
        assert finding.resolved_at is not None
        assert db.query(AuditLog).count() >= 2
    checks.append("08 remediation status update and audit log OK")

reports = sorted(REPORT_DIR.glob("*.json"))
assert len(reports) >= 7, reports
for report in reports:
    json.loads(report.read_text(encoding="utf-8"))
checks.append(f"09 raw scanner report JSON files OK ({len(reports)} generated)")

print("SATRIA END-TO-END SMOKE TEST PASSED")
for item in checks:
    print(f"- {item}")
